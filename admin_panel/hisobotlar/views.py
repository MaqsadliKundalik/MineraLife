from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Sum, F, Count, Avg, Q, Case, When, IntegerField, DecimalField
from django.utils.timezone import now
from django.core.paginator import Paginator
from datetime import timedelta, date
from collections import defaultdict
from orders.models import Order
import json
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def reports_view(request):
    # Foydalanuvchi filterini olish
    today = now().date()
    start = request.GET.get("start")
    end = request.GET.get("end")
    quick = request.GET.get("quick", "month")  # default = hozirgi oy
    page_num = request.GET.get("page", 1)

    if quick == "today":
        start_date = today
        end_date = today
    elif quick == "week":
        # Haftaning boshidan (dushanba) bugungi kungacha
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif quick == "month":
        start_date = today.replace(day=1)
        end_date = today
    elif quick == "6months":
        start_date = today - timedelta(days=180)
        end_date = today
    elif quick == "year":
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        try:
            start_date = date.fromisoformat(start)
            end_date = date.fromisoformat(end)
        except Exception:
            start_date = today.replace(day=1)
            end_date = today

    # Asosiy queryset - select_related bilan optimizatsiya
    qs = Order.objects.filter(
        effective_date__range=[start_date, end_date],
        status="completed"  # faqat tugagan buyurtmalar
    ).select_related("client", "courier")

    # === UMUMIY STATISTIKA ===
    # To'lovlar bo'yicha hisoblash
    cash_total = qs.filter(payment_method="cash").aggregate(total=Sum(F("price") * F("outquantity")))["total"] or 0
    card_total = qs.filter(payment_method="card").aggregate(total=Sum(F("price") * F("outquantity")))["total"] or 0
    pereches_total = qs.filter(payment_method="perechesleniya").aggregate(total=Sum(F("price") * F("outquantity")))["total"] or 0
    debt_total = qs.filter(payment_method="debt").aggregate(total=Sum(F("price") * F("outquantity")))["total"] or 0
    total = cash_total + card_total + pereches_total + debt_total

    # Umumiy buyurtmalar soni va o'rtacha narx
    total_orders = qs.count()
    avg_order_value = qs.aggregate(avg=Avg(F("price") * F("outquantity")))["avg"] or 0
    
    # Jami sotilgan mahsulotlar miqdori
    total_sold = qs.aggregate(total=Sum("outquantity"))["total"] or 0
    total_received = qs.aggregate(total=Sum("inquantity"))["total"] or 0

    # === KUNLIK SAVDO DINAMIKASI (chiziqli diagramma uchun) ===
    daily_sales = defaultdict(lambda: {"revenue": 0, "orders": 0, "quantity": 0})
    
    for order in qs.values("effective_date").annotate(
        revenue=Sum(F("price") * F("outquantity")),
        orders_count=Count("id"),
        quantity=Sum("outquantity")
    ).order_by("effective_date"):
        date_str = order["effective_date"].strftime("%Y-%m-%d")
        daily_sales[date_str] = {
            "revenue": float(order["revenue"] or 0),
            "orders": order["orders_count"],
            "quantity": order["quantity"] or 0
        }
    
    # Sanalarni to'ldirish (bo'sh kunlar uchun)
    current = start_date
    chart_labels = []
    chart_revenue = []
    chart_orders = []
    chart_quantity = []
    
    while current <= end_date:
        date_str = current.strftime("%Y-%m-%d")
        chart_labels.append(date_str)
        chart_revenue.append(daily_sales[date_str]["revenue"])
        chart_orders.append(daily_sales[date_str]["orders"])
        chart_quantity.append(daily_sales[date_str]["quantity"])
        current += timedelta(days=1)

    # === KURYERLAR BO'YICHA STATISTIKA ===
    courier_stats = (
        qs.values("courier__id", "courier__username")
        .annotate(
            total_revenue=Sum(F("price") * F("outquantity")),
            total_orders=Count("id"),
            total_quantity=Sum("outquantity"),
            avg_order=Avg(F("price") * F("outquantity"))
        )
        .order_by("-total_revenue")
    )

    # === MIJOZLAR BO'YICHA STATISTIKA (pagination bilan) ===
    clients_summary = (
        qs.values("client__id", "client__name", "client__caption")
        .annotate(
            total_revenue=Sum(F("price") * F("outquantity")),
            total_orders=Count("id"),
            total_quantity=Sum("outquantity"),
            avg_order=Avg(F("price") * F("outquantity"))
        )
        .order_by("-total_revenue")
    )
    
    # Pagination
    paginator = Paginator(clients_summary, 50)  # Har sahifada 50 ta mijoz
    clients_page = paginator.get_page(page_num)

    # Mijozga qaysi kur'er xizmat qilganini olish (faqat ko'rinayotgan mijozlar uchun)
    client_ids = [c["client__id"] for c in clients_page]
    client_couriers = {}
    for order in qs.filter(client_id__in=client_ids).select_related("client", "courier"):
        if order.client_id not in client_couriers:
            client_couriers[order.client_id] = order.courier

    # === TOP MIJOZLAR (eng ko'p daromad keltirgan 10 ta) ===
    top_clients = clients_summary[:10]

    return render(request, "hisobotlar/reports.html", {
        "start_date": start_date,
        "end_date": end_date,
        "quick": quick,
        
        # Umumiy statistika
        "cash_total": cash_total,
        "card_total": card_total,
        "pereches_total": pereches_total,
        "debt_total": debt_total,
        "total": total,
        "total_orders": total_orders,
        "avg_order_value": avg_order_value,
        "total_sold": total_sold,
        "total_received": total_received,
        
        # Diagramma ma'lumotlari
        "chart_labels": json.dumps(chart_labels),
        "chart_revenue": json.dumps(chart_revenue),
        "chart_orders": json.dumps(chart_orders),
        "chart_quantity": json.dumps(chart_quantity),
        
        # Kuryerlar statistikasi
        "courier_stats": courier_stats,
        
        # Mijozlar (pagination bilan)
        "clients_page": clients_page,
        "client_couriers": client_couriers,
        
        # Top mijozlar
        "top_clients": top_clients,
        "excel_available": EXCEL_AVAILABLE,
    })


def export_excel(request):
    """Excel hisobotini eksport qilish - mijozlar bo'yicha to'lov turlari bilan"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl", status=500)
    
    # Foydalanuvchi filterini olish (reports_view bilan bir xil)
    today = now().date()
    start = request.GET.get("start")
    end = request.GET.get("end")
    quick = request.GET.get("quick", "month")

    if quick == "today":
        start_date = today
        end_date = today
    elif quick == "week":
        # Haftaning boshidan (dushanba) bugungi kungacha
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif quick == "month":
        start_date = today.replace(day=1)
        end_date = today
    elif quick == "6months":
        start_date = today - timedelta(days=180)
        end_date = today
    elif quick == "year":
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        try:
            start_date = date.fromisoformat(start)
            end_date = date.fromisoformat(end)
        except Exception:
            start_date = today.replace(day=1)
            end_date = today

    # Asosiy queryset
    qs = Order.objects.filter(
        effective_date__range=[start_date, end_date],
        status="completed"
    ).select_related("client", "courier")

    # Mijozlar bo'yicha to'lov turlari statistikasi
    clients_data = (
        qs.values("client__id", "client__name", "client__caption")
        .annotate(
            # Naqd
            cash_count=Count("id", filter=Q(payment_method="cash")),
            cash_total=Sum(
                F("price") * F("outquantity"),
                filter=Q(payment_method="cash"),
                output_field=DecimalField()
            ),
            # Karta
            card_count=Count("id", filter=Q(payment_method="card")),
            card_total=Sum(
                F("price") * F("outquantity"),
                filter=Q(payment_method="card"),
                output_field=DecimalField()
            ),
            # Perechesleniya
            pereches_count=Count("id", filter=Q(payment_method="perechesleniya")),
            pereches_total=Sum(
                F("price") * F("outquantity"),
                filter=Q(payment_method="perechesleniya"),
                output_field=DecimalField()
            ),
            # Qarz
            debt_count=Count("id", filter=Q(payment_method="debt")),
            debt_total=Sum(
                F("price") * F("outquantity"),
                filter=Q(payment_method="debt"),
                output_field=DecimalField()
            ),
            # Jami
            total_orders=Count("id"),
            total_revenue=Sum(F("price") * F("outquantity"))
        )
        .order_by("-total_revenue")
    )

    # Excel fayl yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar statistikasi"

    # Stillar
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Sarlavha
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"Mijozlar bo'yicha hisobot ({start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 25

    # Ustun sarlavhalari
    headers = [
        "№",
        "Mijoz nomi",
        "Manzil",
        "Naqd (dona)",
        "Naqd (so'm)",
        "Karta (dona)",
        "Karta (so'm)",
        "Perechesleniya (dona)",
        "Perechesleniya (so'm)",
        "Qarz (dona)",
        "Qarz (so'm)",
        "JAMI (so'm)"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    # Ma'lumotlarni to'ldirish
    row_num = 4
    for idx, client in enumerate(clients_data, 1):
        ws.cell(row=row_num, column=1, value=idx).alignment = center_align
        ws.cell(row=row_num, column=2, value=client["client__name"])
        ws.cell(row=row_num, column=3, value=client["client__caption"] or "—")
        
        # Naqd
        ws.cell(row=row_num, column=4, value=client["cash_count"] or 0).alignment = right_align
        ws.cell(row=row_num, column=5, value=float(client["cash_total"] or 0)).alignment = right_align
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        
        # Karta
        ws.cell(row=row_num, column=6, value=client["card_count"] or 0).alignment = right_align
        ws.cell(row=row_num, column=7, value=float(client["card_total"] or 0)).alignment = right_align
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        
        # Perechesleniya
        ws.cell(row=row_num, column=8, value=client["pereches_count"] or 0).alignment = right_align
        ws.cell(row=row_num, column=9, value=float(client["pereches_total"] or 0)).alignment = right_align
        ws.cell(row=row_num, column=9).number_format = '#,##0'
        
        # Qarz
        ws.cell(row=row_num, column=10, value=client["debt_count"] or 0).alignment = right_align
        ws.cell(row=row_num, column=11, value=float(client["debt_total"] or 0)).alignment = right_align
        ws.cell(row=row_num, column=11).number_format = '#,##0'
        
        # Jami
        total_cell = ws.cell(row=row_num, column=12, value=float(client["total_revenue"] or 0))
        total_cell.alignment = right_align
        total_cell.number_format = '#,##0'
        total_cell.font = Font(bold=True)
        
        # Chegaralar
        for col in range(1, 13):
            ws.cell(row=row_num, column=col).border = border
        
        row_num += 1

    # Ustun kengliklarini sozlash
    column_widths = [5, 30, 25, 12, 15, 12, 15, 18, 20, 12, 15, 18]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Javob
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"hisobot_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
